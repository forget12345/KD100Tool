# -*- encoding: utf-8 -*-
"""
@File    :   miaKd100.py    
@Contact :   timeyuli@163.com


@Modify Time      @Author    @Version    @Desciption
------------      -------    --------    -----------
2019/10/11 16:09   yuli      1.0         None
"""
from datetime import datetime
# 导入字体、边框、颜色以及对齐方式相关库
from openpyxl import Workbook
from openpyxl.styles import Font
import sys, os
from image_rc import *
from time import sleep
from PyQt5.QtWidgets import QMainWindow, QApplication, QFileDialog, QMessageBox, QWidget, QVBoxLayout, QTextBrowser, \
    QHBoxLayout, QPushButton, QMenuBar, QStatusBar, QProgressBar,QSplashScreen
from PyQt5.QtCore import QThread, pyqtSignal, QRect, QMetaObject, QCoreApplication
from PyQt5.QtGui import QIcon,QPixmap


class miaKd100Pricce():
    EXPRESS_INFO = [{"name": "邮政快递包裹", "num": "youzhengguonei", "cnt": 6413328},
                    {"name": "韵达快递", "num": "yunda", "cnt": 5936642},
                    {"name": "中通快递", "num": "zhongtong", "cnt": 3399172},
                    {"name": "顺丰速运", "num": "shunfeng", "cnt": 2748779},
                    {"name": "圆通速递", "num": "yuantong", "cnt": 2742888},
                    {"name": "百世快递", "num": "huitongkuaidi", "cnt": 2558500},
                    {"name": "申通快递", "num": "shentong", "cnt": 2499705}, {"name": "京东物流", "num": "jd", "cnt": 1427612},
                    {"name": "天天快递", "num": "tiantian", "cnt": 865366}, {"name": "EMS", "num": "ems", "cnt": 711984},
                    {"name": "德邦", "num": "debangwuliu", "cnt": 176011},
                    {"name": "优速快递", "num": "youshuwuliu", "cnt": 64342},
                    {"name": "宅急送", "num": "zhaijisong", "cnt": 59478},
                    {"name": "百世快运", "num": "baishiwuliu", "cnt": 54696},
                    {"name": "DHL-中国件", "num": "dhl", "cnt": 44830},
                    {"name": "德邦快递", "num": "debangkuaidi", "cnt": 40452},
                    {"name": "品骏快递", "num": "pjbest", "cnt": 35582},
                    {"name": "TransRush", "num": "transrush", "cnt": 33079},
                    {"name": "安达速递", "num": "adapost", "cnt": 25550},
                    {"name": "中通快运", "num": "zhongtongkuaiyun", "cnt": 25379},
                    {"name": "安能快运", "num": "annengwuliu", "cnt": 22287}, {"name": "盛丰物流", "num": "sfwl", "cnt": 14499},
                    {"name": "USPS", "num": "usps", "cnt": 14357}, {"name": "苏宁物流", "num": "suning", "cnt": 13741},
                    {"name": "EWE全球快递", "num": "ewe", "cnt": 11830},
                    {"name": "邮政标准快递", "num": "youzhengbk", "cnt": 11145},
                    {"name": "EMS-国际件", "num": "emsguoji", "cnt": 10706},
                    {"name": "丹鸟", "num": "danniao", "cnt": 10257},
                    {"name": "FedEx-国际件", "num": "fedex", "cnt": 9727}, {"name": "速尔快递", "num": "suer", "cnt": 9349},
                    {"name": "韵达快运", "num": "yundakuaiyun", "cnt": 9195},
                    {"name": "UPS-全球件", "num": "upsen", "cnt": 9066},
                    {"name": "泛远国际物流", "num": "farlogistis", "cnt": 8356},
                    {"name": "威盛快递", "num": "wherexpess", "cnt": 8011},
                    {"name": "转运四方", "num": "zhuanyunsifang", "cnt": 7685},
                    {"name": "国际包裹", "num": "youzhengguoji", "cnt": 7641},
                    {"name": "国通快递", "num": "guotongkuaidi", "cnt": 7464},
                    {"name": "京广速递", "num": "jinguangsudikuaijian", "cnt": 6706},
                    {"name": "日本（Japan Post）", "num": "japanposten", "cnt": 6696},
                    {"name": "壹米滴答", "num": "yimidida", "cnt": 6436}, {"name": "速必达", "num": "subida", "cnt": 6309},
                    {"name": "芝麻开门", "num": "zhimakaimen", "cnt": 6122},
                    {"name": "联昊通", "num": "lianhaowuliu", "cnt": 6117}, {"name": "跨越速运", "num": "kuayue", "cnt": 5693},
                    {"name": "澳邮中国快运", "num": "auexpress", "cnt": 5610},
                    {"name": "圆通快运", "num": "yuantongkuaiyun", "cnt": 5447},
                    {"name": "联邦快递", "num": "lianbangkuaidi", "cnt": 5317},
                    {"name": "D速快递", "num": "dsukuaidi", "cnt": 5025}, {"name": "DHL-全球件", "num": "dhlen", "cnt": 4945},
                    {"name": "TNT", "num": "tnt", "cnt": 4669}, {"name": "安得物流", "num": "annto", "cnt": 4658},
                    {"name": "宇鑫物流", "num": "yuxinwuliu", "cnt": 4629},
                    {"name": "中粮鲜到家物流", "num": "zlxdjwl", "cnt": 4476},
                    {"name": "美快国际物流", "num": "meiquick", "cnt": 4439}, {"name": "程光快递", "num": "flyway", "cnt": 4372},
                    {"name": "九曳供应链", "num": "jiuyescm", "cnt": 4314},
                    {"name": "龙邦速递", "num": "longbanwuliu", "cnt": 4045},
                    {"name": "全峰快递", "num": "quanfengkuaidi", "cnt": 3767},
                    {"name": "日日顺物流", "num": "rrs", "cnt": 3551},
                    {"name": "富腾达国际货运", "num": "ftd", "cnt": 3395}, {"name": "邦泰快运", "num": "btexpress", "cnt": 3246},
                    {"name": "方舟速递", "num": "arkexpress", "cnt": 3147}, {"name": "联合快递", "num": "gslhkd", "cnt": 3089},
                    {"name": "Xlobo贝海国际", "num": "xlobo", "cnt": 3052}, {"name": "特急送", "num": "lntjs", "cnt": 2911},
                    {"name": "万象物流", "num": "wanxiangwuliu", "cnt": 2897},
                    {"name": "EMS包裹", "num": "emsbg", "cnt": 2833},
                    {"name": "洋包裹", "num": "yangbaoguo", "cnt": 2676}, {"name": "UPS", "num": "ups", "cnt": 2642},
                    {"name": "佳吉快运", "num": "jiajiwuliu", "cnt": 2547},
                    {"name": "C&C国际速递", "num": "cncexp", "cnt": 2501},
                    {"name": "天地华宇", "num": "tiandihuayu", "cnt": 2452},
                    {"name": "FedEx-美国件", "num": "fedexus", "cnt": 2436}, {"name": "斑马物流", "num": "banma", "cnt": 2415},
                    {"name": "澳洲飞跃物流", "num": "rlgaus", "cnt": 2325},
                    {"name": "中速快递", "num": "zhongsukuaidi", "cnt": 2319},
                    {"name": "优邦速运", "num": "ubonex", "cnt": 2297},
                    {"name": "蓝天快递", "num": "blueskyexpress", "cnt": 2293},
                    {"name": "Superb Grace", "num": "superb", "cnt": 2195},
                    {"name": "中通国际", "num": "zhongtongguoji", "cnt": 2121},
                    {"name": "亚风速递", "num": "yafengsudi", "cnt": 2053},
                    {"name": "加运美", "num": "jiayunmeiwuliu", "cnt": 2038},
                    {"name": "叮咚澳洲转运", "num": "dindon", "cnt": 1904},
                    {"name": "安世通快递", "num": "astexpress", "cnt": 1839},
                    {"name": "长江国际速递", "num": "changjiang", "cnt": 1804},
                    {"name": "DHL-德国件（DHL Deutschland）", "num": "dhlde", "cnt": 1728},
                    {"name": "COE", "num": "coe", "cnt": 1715}, {"name": "迅达速递", "num": "xdexpress", "cnt": 1665},
                    {"name": "同城快寄", "num": "shpost", "cnt": 1632}, {"name": "中翼国际物流", "num": "chnexp", "cnt": 1605},
                    {"name": "比利时（Bpost）", "num": "bpost", "cnt": 1570},
                    {"name": "承诺达", "num": "ytchengnuoda", "cnt": 1542},
                    {"name": "极地快递", "num": "polarexpress", "cnt": 1503},
                    {"name": "美通", "num": "valueway", "cnt": 1439},
                    {"name": "百事亨通", "num": "bsht", "cnt": 1364}, {"name": "OnTrac", "num": "ontrac", "cnt": 1217},
                    {"name": "中外运", "num": "esinotrans", "cnt": 1188}, {"name": "Aramex", "num": "aramex", "cnt": 1182},
                    {"name": "信丰物流", "num": "xinfengwuliu", "cnt": 1150},
                    {"name": "新顺丰（NSF）", "num": "nsf", "cnt": 1134},
                    {"name": "微特派", "num": "weitepai", "cnt": 1126},
                    {"name": "中国香港(HongKong Post)", "num": "hkpost", "cnt": 1107},
                    {"name": "皇家物流", "num": "pfcexpress", "cnt": 1085},
                    {"name": "中环快递", "num": "zhonghuan", "cnt": 1083},
                    {"name": "欧亚专线", "num": "euasia", "cnt": 1073}, {"name": "百世云配", "num": "baishiyp", "cnt": 1015},
                    {"name": "苏通快运", "num": "zjstky", "cnt": 1014}, {"name": "光线速递", "num": "gxwl", "cnt": 1003},
                    {"name": "中远e环球", "num": "cosco", "cnt": 988}, {"name": "速达通", "num": "sdto", "cnt": 975},
                    {"name": "美西快递", "num": "meixi", "cnt": 956}, {"name": "YUN TRACK", "num": "yuntrack", "cnt": 930},
                    {"name": "TNT UK", "num": "tntuk", "cnt": 916}, {"name": "泛捷国际速递", "num": "epanex", "cnt": 879},
                    {"name": "保加利亚（Bulgarian Posts）", "num": "bulgarian", "cnt": 837},
                    {"name": "景顺物流", "num": "jingshun", "cnt": 832}, {"name": "易达通快递", "num": "qexpress", "cnt": 832},
                    {"name": "黄马甲", "num": "huangmajia", "cnt": 823}, {"name": "港快速递", "num": "gdkd", "cnt": 791},
                    {"name": "联合速运", "num": "unitedex", "cnt": 776},
                    {"name": "EMS-国际件-英文", "num": "emsinten", "cnt": 772},
                    {"name": "UEQ快递", "num": "ueq", "cnt": 769}, {"name": "宅急便", "num": "zhaijibian", "cnt": 769},
                    {"name": "广东邮政", "num": "guangdongyouzhengwuliu", "cnt": 746},
                    {"name": "加拿大(Canada Post)", "num": "canpost", "cnt": 742},
                    {"name": "万家康物流", "num": "wjkwl", "cnt": 713}, {"name": "顺心捷达", "num": "sxjdfreight", "cnt": 699},
                    {"name": "新杰物流", "num": "sunjex", "cnt": 678}, {"name": "西翼物流", "num": "westwing", "cnt": 662},
                    {"name": "荷兰邮政(PostNL international registered mail)", "num": "postnl", "cnt": 658},
                    {"name": "澳德物流", "num": "auod", "cnt": 621}, {"name": "DPD UK", "num": "dpduk", "cnt": 585},
                    {"name": "顺达快递", "num": "sundarexpress", "cnt": 573},
                    {"name": "运通速运", "num": "yuntong", "cnt": 569},
                    {"name": "瑞士邮政", "num": "swisspostcn", "cnt": 557}, {"name": "EMS-英文", "num": "emsen", "cnt": 548},
                    {"name": "Fedex-国际件-中文", "num": "fedexcn", "cnt": 530},
                    {"name": "优速通达", "num": "yousutongda", "cnt": 506}, {"name": "环球速运", "num": "huanqiu", "cnt": 498},
                    {"name": "安鲜达", "num": "exfresh", "cnt": 483}, {"name": "疯狂快递", "num": "crazyexpress", "cnt": 481},
                    {"name": "优优速递", "num": "youyou", "cnt": 474}, {"name": "TNT-全球件", "num": "tnten", "cnt": 434},
                    {"name": "SYNSHIP快递", "num": "synship", "cnt": 427},
                    {"name": "Hermes", "num": "hermes", "cnt": 425},
                    {"name": "运通中港", "num": "yuntongkuaidi", "cnt": 420}, {"name": "玥玛速运", "num": "yue777", "cnt": 405},
                    {"name": "皮牙子快递", "num": "bazirim", "cnt": 404}, {"name": "百通物流", "num": "buytong", "cnt": 404},
                    {"name": "e直运", "num": "edtexpress", "cnt": 402}, {"name": "三象速递", "num": "sxexpress", "cnt": 400},
                    {"name": "优联吉运", "num": "uluckex", "cnt": 392}, {"name": "易客满", "num": "ecmscn", "cnt": 386},
                    {"name": "明达国际速递", "num": "tmwexpress", "cnt": 378},
                    {"name": "速递中国", "num": "sendtochina", "cnt": 377},
                    {"name": "华夏国际速递", "num": "uschuaxia", "cnt": 372},
                    {"name": "全一快递", "num": "quanyikuaidi", "cnt": 368},
                    {"name": "盛辉物流", "num": "shenghuiwuliu", "cnt": 364}, {"name": "安迅物流", "num": "anxl", "cnt": 362},
                    {"name": "恒路物流", "num": "hengluwuliu", "cnt": 355}, {"name": "DPEX", "num": "dpex", "cnt": 354},
                    {"name": "新加坡小包(Singapore Post)", "num": "singpost", "cnt": 347},
                    {"name": "中邮物流", "num": "zhongyouwuliu", "cnt": 337},
                    {"name": "英国小包（Royal Mail）", "num": "royalmail", "cnt": 330},
                    {"name": "邮邦国际", "num": "youban", "cnt": 305}, {"name": "南方传媒物流", "num": "ndwl", "cnt": 297},
                    {"name": "中铁快运", "num": "ztky", "cnt": 296}, {"name": "合众速递(UCS）", "num": "ucs", "cnt": 282},
                    {"name": "AAE-中国件", "num": "aae", "cnt": 281}, {"name": "TST速运通", "num": "tstexp", "cnt": 272},
                    {"name": "CNPEX中邮快递", "num": "cnpex", "cnt": 269}, {"name": "宏递快运", "num": "hd", "cnt": 268},
                    {"name": "大马鹿", "num": "idamalu", "cnt": 262}, {"name": "Toll", "num": "dpexen", "cnt": 260},
                    {"name": "中铁物流", "num": "zhongtiewuliu", "cnt": 255}, {"name": "龙邦物流", "num": "lbex", "cnt": 244},
                    {"name": "DPD", "num": "dpd", "cnt": 238}, {"name": "如家国际快递", "num": "homecourier", "cnt": 235},
                    {"name": "晟邦物流", "num": "nanjingshengbang", "cnt": 233},
                    {"name": "递四方", "num": "disifang", "cnt": 222},
                    {"name": "程光快递", "num": "chengguangkuaidi", "cnt": 219},
                    {"name": "众川国际", "num": "zhongchuan", "cnt": 217},
                    {"name": "速派快递(FastGo)", "num": "fastgo", "cnt": 212},
                    {"name": "快捷速递", "num": "kuaijiesudi", "cnt": 205},
                    {"name": "CHS中环国际快递", "num": "chszhonghuanguoji", "cnt": 199},
                    {"name": "美国快递", "num": "meiguokuaidi", "cnt": 198},
                    {"name": "英国大包、EMS（Parcel Force）", "num": "parcelforce", "cnt": 197},
                    {"name": "OCS", "num": "ocs", "cnt": 194}, {"name": "澳天速运", "num": "aotsd", "cnt": 193},
                    {"name": "可可树美中速运", "num": "excocotree", "cnt": 191},
                    {"name": "顺丰-美国件", "num": "shunfengen", "cnt": 185},
                    {"name": "韩国邮政", "num": "koreapostcn", "cnt": 184},
                    {"name": "如风达", "num": "rufengda", "cnt": 184},
                    {"name": "澳大利亚(Australia Post)", "num": "auspost", "cnt": 183},
                    {"name": "澳世速递", "num": "ausexpress", "cnt": 182}, {"name": "锦程快递", "num": "hrex", "cnt": 171},
                    {"name": "飞洋快递", "num": "shipgce", "cnt": 170}, {"name": "日本郵便", "num": "japanpost", "cnt": 167},
                    {"name": "安能快递", "num": "ane66", "cnt": 164}, {"name": "春风物流", "num": "spring56", "cnt": 161},
                    {"name": "恒宇运通", "num": "hyytes", "cnt": 160},
                    {"name": "中通（带电话）", "num": "zhongtongphone", "cnt": 160},
                    {"name": "华中快递", "num": "cpsair", "cnt": 157},
                    {"name": "中国澳门(Macau Post)", "num": "macao", "cnt": 157},
                    {"name": "法国(La Poste)", "num": "csuivi", "cnt": 151}, {"name": "远成快运", "num": "ycgky", "cnt": 151},
                    {"name": "飞云快递系统", "num": "fyex", "cnt": 146}, {"name": "速派快递", "num": "fastgoexpress", "cnt": 144},
                    {"name": "嘉诚速达", "num": "jcsuda", "cnt": 143}, {"name": "速通物流", "num": "sut56", "cnt": 141},
                    {"name": "增速跨境 ", "num": "zyzoom", "cnt": 141}, {"name": "猛犸速递", "num": "mmlogi", "cnt": 139},
                    {"name": "赛澳递for买卖宝", "num": "saiaodimmb", "cnt": 137},
                    {"name": "ZTE中兴物流", "num": "zteexpress", "cnt": 137}, {"name": "星空国际", "num": "wlwex", "cnt": 134},
                    {"name": "飞邦快递", "num": "fbkd", "cnt": 133}, {"name": "威时沛运货运", "num": "wtdchina", "cnt": 133},
                    {"name": "汇通天下物流", "num": "httx56", "cnt": 132},
                    {"name": "日日顺智慧物联", "num": "gooday365", "cnt": 129},
                    {"name": "货运皇", "num": "kingfreight", "cnt": 123}, {"name": "天马迅达", "num": "tianma", "cnt": 121},
                    {"name": "万家物流", "num": "wanjiawuliu", "cnt": 119},
                    {"name": "成都东骏物流", "num": "dongjun", "cnt": 115},
                    {"name": "嘉里大通", "num": "jialidatong", "cnt": 115}, {"name": "捷安达", "num": "jieanda", "cnt": 114},
                    {"name": "一号线", "num": "lineone", "cnt": 113},
                    {"name": "盛丰物流", "num": "shengfengwuliu", "cnt": 113},
                    {"name": "韩国（Korea Post）", "num": "koreapost", "cnt": 111},
                    {"name": "大田物流", "num": "datianwuliu", "cnt": 110}, {"name": "申通快运", "num": "stoe56", "cnt": 108},
                    {"name": "意大利(Poste Italiane)", "num": "italiane", "cnt": 105},
                    {"name": "商海德物流", "num": "shd56", "cnt": 101}, {"name": "西游寄", "num": "xiyoug", "cnt": 101},
                    {"name": "天翼快递", "num": "tykd", "cnt": 100},
                    {"name": "荷兰邮政-中文(PostNL international registered mail)", "num": "postnlcn", "cnt": 96},
                    {"name": "法国小包（colissimo）", "num": "colissimo", "cnt": 94},
                    {"name": "新元国际", "num": "xynyc", "cnt": 94},
                    {"name": "锋鸟物流", "num": "beebird", "cnt": 92}, {"name": "尚橙物流", "num": "shangcheng", "cnt": 91},
                    {"name": "迅速快递", "num": "xunsuexpress", "cnt": 91}, {"name": "秦邦快运", "num": "qbexpress", "cnt": 90},
                    {"name": "TNT Australia", "num": "tntau", "cnt": 90},
                    {"name": "新西兰（New Zealand Post）", "num": "newzealand", "cnt": 87},
                    {"name": "中环转运", "num": "zhonghuanus", "cnt": 85},
                    {"name": "英国邮政小包", "num": "royalmailcn", "cnt": 83},
                    {"name": "鸿泰物流", "num": "hnht56", "cnt": 82},
                    {"name": "德国(Deutsche Post)", "num": "deutschepost", "cnt": 81},
                    {"name": "丰程物流", "num": "sccod", "cnt": 81}, {"name": "深圳德创物流", "num": "dechuangwuliu", "cnt": 79},
                    {"name": "City-Link", "num": "citylink", "cnt": 78}, {"name": "蜜蜂速递", "num": "bee001", "cnt": 77},
                    {"name": "全联速运", "num": "guexp", "cnt": 76}, {"name": "海带宝", "num": "haidaibao", "cnt": 74},
                    {"name": "佐川急便", "num": "sagawa", "cnt": 74}, {"name": "昌宇国际", "num": "changwooair", "cnt": 71},
                    {"name": "商桥物流", "num": "shangqiao56", "cnt": 71}, {"name": "天翼物流", "num": "tywl99", "cnt": 70},
                    {"name": "卓志速运", "num": "chinaicip", "cnt": 69}, {"name": "景光物流", "num": "jgwl", "cnt": 69},
                    {"name": "铁中快运", "num": "tzky", "cnt": 68},
                    {"name": "西班牙(Correos de Espa?a)", "num": "correosdees", "cnt": 67},
                    {"name": "加运美速递", "num": "jym56", "cnt": 67}, {"name": "嘉里大荣物流", "num": "kerrytj", "cnt": 67},
                    {"name": "平安达腾飞", "num": "pingandatengfei", "cnt": 67},
                    {"name": "TRAKPAK", "num": "trakpak", "cnt": 67}, {"name": "一运全成物流", "num": "yyqc56", "cnt": 66},
                    {"name": "中骅物流", "num": "chunghwa56", "cnt": 63}, {"name": "亚洲顺物流", "num": "yzswuliu", "cnt": 63},
                    {"name": "鼎润物流", "num": "la911", "cnt": 62}, {"name": "中运全速", "num": "topspeedex", "cnt": 62},
                    {"name": "台湾（中华邮政）", "num": "postserv", "cnt": 61}, {"name": "考拉速递", "num": "koalaexp", "cnt": 59},
                    {"name": "英超物流", "num": "yingchao", "cnt": 59}, {"name": "青岛安捷快递", "num": "anjiekuaidi", "cnt": 58},
                    {"name": "蓝天快递", "num": "lantiankuaidi", "cnt": 57},
                    {"name": "EFSPOST", "num": "efspost", "cnt": 56},
                    {"name": "速通物流", "num": "sutongwuliu", "cnt": 55},
                    {"name": "瑞士(Swiss Post)", "num": "swisspost", "cnt": 55},
                    {"name": "深圳邮政", "num": "szyouzheng", "cnt": 54}, {"name": "星云速递", "num": "nebuex", "cnt": 53},
                    {"name": "一速递", "num": "oneexpress", "cnt": 51},
                    {"name": "瑞典（Sweden Post）", "num": "ruidianyouzheng", "cnt": 51},
                    {"name": "堡昕德速递", "num": "bosind", "cnt": 50},
                    {"name": "陆本速递 LUBEN EXPRESS", "num": "luben", "cnt": 50},
                    {"name": "家家通快递", "num": "newsway", "cnt": 48}, {"name": "一智通", "num": "1ziton", "cnt": 46},
                    {"name": "中集冷云", "num": "cccc58", "cnt": 43}, {"name": "佳怡物流", "num": "jiayiwuliu", "cnt": 43},
                    {"name": "民航快递", "num": "minghangkuaidi", "cnt": 43}, {"name": "易达通", "num": "yidatong", "cnt": 43},
                    {"name": "EFS Post（平安快递）", "num": "efs", "cnt": 42},
                    {"name": "成都立即送", "num": "lijisong", "cnt": 42},
                    {"name": "一号仓", "num": "onehcang", "cnt": 40}, {"name": "易境达国际物流", "num": "uscbexpress", "cnt": 40},
                    {"name": "红马甲物流", "num": "sxhongmajia", "cnt": 38}, {"name": "黑猫宅急便", "num": "tcat", "cnt": 37},
                    {"name": "新邦物流", "num": "xinbangwuliu", "cnt": 36}, {"name": "海淘物流", "num": "ht22", "cnt": 35},
                    {"name": "泰国中通CTO", "num": "ctoexp", "cnt": 34}, {"name": "金岸物流", "num": "jinan", "cnt": 33},
                    {"name": "新速航", "num": "sunspeedy", "cnt": 33}, {"name": "泰进物流", "num": "taijin", "cnt": 33},
                    {"name": "安信达", "num": "anxindakuaixi", "cnt": 32},
                    {"name": "荷兰速递(Nederland Post)", "num": "nederlandpost", "cnt": 32},
                    {"name": "申通国际", "num": "stosolution", "cnt": 32},
                    {"name": "法国大包、EMS-法文（Chronopost France）", "num": "chronopostfra", "cnt": 31},
                    {"name": "瀚朝物流", "num": "hac56", "cnt": 31}, {"name": "顺通快递", "num": "stkd", "cnt": 31},
                    {"name": "EMS物流", "num": "emswuliu", "cnt": 30}, {"name": "TCXB国际物流", "num": "tcxbthai", "cnt": 30},
                    {"name": "百腾物流", "num": "baitengwuliu", "cnt": 29}, {"name": "秦远物流", "num": "qinyuan", "cnt": 29},
                    {"name": "优海国际速递", "num": "uhi", "cnt": 29}, {"name": "远成物流", "num": "yuanchengwuliu", "cnt": 29},
                    {"name": "贝业物流", "num": "boyol", "cnt": 28}, {"name": "能达速递", "num": "ganzhongnengda", "cnt": 28},
                    {"name": "印度(India Post)", "num": "india", "cnt": 28},
                    {"name": "LaserShip", "num": "lasership", "cnt": 28},
                    {"name": "顺丰-繁体", "num": "shunfenghk", "cnt": 28},
                    {"name": "E通速递", "num": "etong", "cnt": 27}, {"name": "丰通快运", "num": "ftky365", "cnt": 27},
                    {"name": "环国运物流", "num": "hgy56", "cnt": 27},
                    {"name": "新加坡EMS、大包(Singapore Speedpost)", "num": "speedpost", "cnt": 27},
                    {"name": "八达通", "num": "bdatong", "cnt": 26}, {"name": "美国申通", "num": "stoexpress", "cnt": 26},
                    {"name": "城铁速递", "num": "cex", "cnt": 25}, {"name": "东西E全运", "num": "ecotransite", "cnt": 25},
                    {"name": "飞豹快递", "num": "feibaokuaidi", "cnt": 23}, {"name": "启辰国际速递", "num": "qichen", "cnt": 23},
                    {"name": "广东速腾物流", "num": "suteng", "cnt": 23},
                    {"name": "澳速通国际速递", "num": "jetexpressgroup", "cnt": 22},
                    {"name": "尼尔快递", "num": "nell", "cnt": 22},
                    {"name": "俄罗斯邮政(Russian Post)", "num": "pochta", "cnt": 22},
                    {"name": "德坤物流", "num": "dekuncn", "cnt": 21}, {"name": "YODEL", "num": "yodel", "cnt": 21},
                    {"name": "爱尔兰(An Post)", "num": "anposten", "cnt": 20},
                    {"name": "DHL Benelux", "num": "dhlbenelux", "cnt": 20},
                    {"name": "黑猫速运", "num": "heimao56", "cnt": 20},
                    {"name": "蓝天国际快递", "num": "ltx", "cnt": 19}, {"name": "海红网送", "num": "haihongwangsong", "cnt": 18},
                    {"name": "鹏远国际速递", "num": "pengyuanexpress", "cnt": 17},
                    {"name": "科捷物流", "num": "kejie", "cnt": 16},
                    {"name": "全球快运", "num": "abcglobal", "cnt": 15}, {"name": "久久物流", "num": "jiujiuwl", "cnt": 15},
                    {"name": "柬埔寨中通", "num": "khzto", "cnt": 15}, {"name": "腾达速递", "num": "nntengda", "cnt": 15},
                    {"name": "泰国138国际物流", "num": "sd138", "cnt": 15}, {"name": "中澳速递", "num": "cnausu", "cnt": 14},
                    {"name": "海联快递", "num": "hltop", "cnt": 14},
                    {"name": "马来西亚小包（Malaysia Post(Registered)）", "num": "malaysiapost", "cnt": 14},
                    {"name": "增益速递", "num": "zengyisudi", "cnt": 14},
                    {"name": "比利时国际(Bpost international)", "num": "bpostinter", "cnt": 13},
                    {"name": "诚和通", "num": "cht361", "cnt": 13}, {"name": "FQ狂派速递", "num": "freakyquick", "cnt": 13},
                    {"name": "集先锋快递", "num": "jxfex", "cnt": 13}, {"name": "MyHermes", "num": "myhermes", "cnt": 13},
                    {"name": "泰国（Thailand Thai Post）", "num": "thailand", "cnt": 13},
                    {"name": "美国云达", "num": "yundaexus", "cnt": 13},
                    {"name": "EASY EXPRESS", "num": "easyexpress", "cnt": 12},
                    {"name": "宇佳物流", "num": "yujiawl", "cnt": 12},
                    {"name": "airpak expresss", "num": "airpak", "cnt": 11},
                    {"name": "dhl小包", "num": "dhlecommerce", "cnt": 11},
                    {"name": "格鲁吉亚(Georgian Pos）", "num": "georgianpost", "cnt": 11},
                    {"name": "德尚国际速递", "num": "gslexpress", "cnt": 11},
                    {"name": "捷特快递", "num": "jietekuaidi", "cnt": 11},
                    {"name": "转运中国", "num": "uszcn", "cnt": 11}, {"name": "香港伟豪国际物流", "num": "whgjkd", "cnt": 11},
                    {"name": "全日通", "num": "quanritongkuaidi", "cnt": 10},
                    {"name": "track-parcel", "num": "trackparcel", "cnt": 9},
                    {"name": "五六快运", "num": "wuliuky", "cnt": 9},
                    {"name": "纵通速运", "num": "ynztsy", "cnt": 9}, {"name": "永昌物流", "num": "yongchangwuliu", "cnt": 9},
                    {"name": "运通中港快递", "num": "ytkd", "cnt": 9}, {"name": "佰麒快递", "num": "beckygo", "cnt": 8},
                    {"name": "大洋物流", "num": "dayangwuliu", "cnt": 8}, {"name": "泰国中通ZTO", "num": "thaizto", "cnt": 8},
                    {"name": "中邮速递", "num": "wondersyd", "cnt": 8}, {"name": "传喜物流", "num": "chuanxiwuliu", "cnt": 7},
                    {"name": "中国香港(HongKong Post)英文", "num": "hkposten", "cnt": 7},
                    {"name": "UPS i-parcel", "num": "iparcel", "cnt": 7}, {"name": "乐天速递", "num": "ltexp", "cnt": 7},
                    {"name": "澳洲迈速快递", "num": "maxeedexpress", "cnt": 7},
                    {"name": "PCA Express", "num": "pcaexpress", "cnt": 7}, {"name": "全速物流", "num": "quansu", "cnt": 7},
                    {"name": "云达通", "num": "ydglobe", "cnt": 7}, {"name": "圆通国际", "num": "yuantongguoji", "cnt": 7},
                    {"name": "中天万运", "num": "zhongtianwanyun", "cnt": 7}, {"name": "安捷物流", "num": "anjie88", "cnt": 6},
                    {"name": "邦送物流", "num": "bangsongwuliu", "cnt": 6}, {"name": "河北橙配", "num": "chengpei", "cnt": 6},
                    {"name": "CNAIR", "num": "cnair", "cnt": 6}, {"name": "FOX国际快递", "num": "fox", "cnt": 6},
                    {"name": "好来运", "num": "hlyex", "cnt": 6}, {"name": "木春货运", "num": "mchy", "cnt": 6},
                    {"name": "沙特阿拉伯(Saudi Post)", "num": "saudipost", "cnt": 6},
                    {"name": "联运通物流", "num": "szuem", "cnt": 6}, {"name": "株式会社T.M.G", "num": "tmg", "cnt": 6},
                    {"name": "新宁物流", "num": "xinning", "cnt": 6}, {"name": "源安达", "num": "yuananda", "cnt": 6},
                    {"name": "百福东方", "num": "baifudongfang", "cnt": 5},
                    {"name": "Fastway Ireland", "num": "fastway", "cnt": 5},
                    {"name": "芬兰(Itella Posti Oy)", "num": "finland", "cnt": 5},
                    {"name": "GLS", "num": "gls", "cnt": 5},
                    {"name": "高捷快运", "num": "goldjet", "cnt": 5}, {"name": "驿扬国际速运", "num": "iyoungspeed", "cnt": 5},
                    {"name": "Landmark Global", "num": "landmarkglobal", "cnt": 5},
                    {"name": "豌豆物流", "num": "wandougongzhu", "cnt": 5},
                    {"name": "祥龙运通物流", "num": "xianglongyuntong", "cnt": 5},
                    {"name": "易欧洲国际物流", "num": "yiouzhou", "cnt": 5}, {"name": "德国雄鹰速递", "num": "adlerlogi", "cnt": 4},
                    {"name": "北京EMS", "num": "bjemstckj", "cnt": 4},
                    {"name": "捷克（?eská po?ta）", "num": "ceskaposta", "cnt": 4},
                    {"name": "中邮电商", "num": "chinapostcb", "cnt": 4}, {"name": "duodao56", "num": "duodao56", "cnt": 4},
                    {"name": "天天快物流", "num": "guoeryue", "cnt": 4},
                    {"name": "爱尔兰(An Post)", "num": "ireland", "cnt": 4},
                    {"name": "佳成快递 ", "num": "jiacheng", "cnt": 4}, {"name": "急先达", "num": "jixianda", "cnt": 4},
                    {"name": "快服务", "num": "kfwnet", "cnt": 4}, {"name": "華信物流WTO", "num": "logistics", "cnt": 4},
                    {"name": "6LS EXPRESS", "num": "lsexpress", "cnt": 4},
                    {"name": "马来西亚大包、EMS（Malaysia Post(parcel,EMS)）", "num": "malaysiaems", "cnt": 4},
                    {"name": "荷兰包裹(PostNL International Parcels)", "num": "postnlpacle", "cnt": 4},
                    {"name": "7E速递", "num": "qesd", "cnt": 4}, {"name": "圣安物流", "num": "shenganwuliu", "cnt": 4},
                    {"name": "签收快递", "num": "signedexpress", "cnt": 4}, {"name": "淘布斯国际物流", "num": "taoplus", "cnt": 4},
                    {"name": "易通达", "num": "yitongda", "cnt": 4}, {"name": "安达信", "num": "advancing", "cnt": 3},
                    {"name": "捷记方舟", "num": "ajexpress", "cnt": 3}, {"name": "新干线快递", "num": "anlexpress", "cnt": 3},
                    {"name": "Asendia USA", "num": "asendiausa", "cnt": 3},
                    {"name": "中联速递", "num": "auvanda", "cnt": 3},
                    {"name": "比利时(Belgium Post)", "num": "belgiumpost", "cnt": 3},
                    {"name": "奔腾物流", "num": "benteng", "cnt": 3}, {"name": "百千诚物流", "num": "bqcwl", "cnt": 3},
                    {"name": "河南次晨达", "num": "ccd", "cnt": 3}, {"name": "创一快递", "num": "chuangyi", "cnt": 3},
                    {"name": "邦通国际", "num": "comexpress", "cnt": 3}, {"name": "大道物流", "num": "dadaoex", "cnt": 3},
                    {"name": "递达速运", "num": "didasuyun", "cnt": 3}, {"name": "CJ物流", "num": "doortodoor", "cnt": 3},
                    {"name": "加州猫速递", "num": "jiazhoumao", "cnt": 3},
                    {"name": "骏丰国际速递", "num": "junfengguoji", "cnt": 3},
                    {"name": "卢森堡(Luxembourg Post)", "num": "luxembourg", "cnt": 3},
                    {"name": "新亚物流", "num": "nalexpress", "cnt": 3}, {"name": "诚一物流", "num": "parcelchina", "cnt": 3},
                    {"name": "坦桑尼亚（Tanzania Posts Corporation）", "num": "posta", "cnt": 3},
                    {"name": "荷兰邮政-中国件", "num": "postnlchina", "cnt": 3}, {"name": "全通快运", "num": "quantwl", "cnt": 3},
                    {"name": "三态速递", "num": "santaisudi", "cnt": 3}, {"name": "SHL畅灵国际物流", "num": "shlexp", "cnt": 3},
                    {"name": "顺捷达", "num": "shunjieda", "cnt": 3}, {"name": "顺捷丰达", "num": "shunjiefengda", "cnt": 3},
                    {"name": "USPSCN", "num": "uspscn", "cnt": 3},
                    {"name": "乌兹别克斯坦(Post of Uzbekistan)", "num": "uzbekistan", "cnt": 3},
                    {"name": "越南EMS(VNPost Express)", "num": "vnpost", "cnt": 3},
                    {"name": "沃埃家", "num": "wowvip", "cnt": 3}, {"name": "一起送", "num": "yiqisong", "cnt": 3},
                    {"name": "振捷国际货运", "num": "zjgj56", "cnt": 3},
                    {"name": "阿尔巴尼亚(Posta shqipatre)", "num": "albania", "cnt": 2},
                    {"name": "心怡物流", "num": "alog", "cnt": 2},
                    {"name": "奥地利(Austrian Post)", "num": "austria", "cnt": 2},
                    {"name": "帮帮发", "num": "bangbangpost", "cnt": 2}, {"name": "BCWELT", "num": "bcwelt", "cnt": 2},
                    {"name": "BlueDart", "num": "bluedart", "cnt": 2},
                    {"name": "巴西(Brazil Post/Correios)", "num": "brazilposten", "cnt": 2},
                    {"name": "出口易", "num": "chukou1", "cnt": 2}, {"name": "中国香港骏辉物流", "num": "chunfai", "cnt": 2},
                    {"name": "DCS", "num": "dcs", "cnt": 2},
                    {"name": "波兰小包(Poczta Polska)", "num": "emonitoring", "cnt": 2},
                    {"name": "europeanecom", "num": "europeanecom", "cnt": 2},
                    {"name": "E速达", "num": "exsuda", "cnt": 2},
                    {"name": "凡宇快递", "num": "fanyukuaidi", "cnt": 2},
                    {"name": "加拿大联通快运", "num": "fastontime", "cnt": 2},
                    {"name": "FedEx-英国件（FedEx UK)", "num": "fedexuk", "cnt": 2},
                    {"name": "飞康达", "num": "feikangda", "cnt": 2}, {"name": "共速达", "num": "gongsuda", "cnt": 2},
                    {"name": "合心速递", "num": "hexinexpress", "cnt": 2}, {"name": "猴急送", "num": "hjs", "cnt": 2},
                    {"name": "鸿讯物流", "num": "hongxun", "cnt": 2}, {"name": "中国香港环球快运", "num": "huanqiuabc", "cnt": 2},
                    {"name": "匈牙利（Magyar Posta）", "num": "hungary", "cnt": 2},
                    {"name": "冰岛(Iceland Post)", "num": "iceland", "cnt": 2},
                    {"name": "骏达快递", "num": "jdexpressusa", "cnt": 2}, {"name": "泽西岛", "num": "jerseypost", "cnt": 2},
                    {"name": "捷邦物流", "num": "jieborne", "cnt": 2}, {"name": "快速递", "num": "ksudi", "cnt": 2},
                    {"name": "鲁通快运", "num": "lutong", "cnt": 2},
                    {"name": "毛里求斯(Mauritius Post)", "num": "mauritius", "cnt": 2},
                    {"name": "昂威物流", "num": "onway", "cnt": 2}, {"name": "PostElbe", "num": "postelbe", "cnt": 2},
                    {"name": "叙利亚(Syrian Post)", "num": "republic", "cnt": 2},
                    {"name": "日日顺快线", "num": "rrskx", "cnt": 2},
                    {"name": "S2C", "num": "s2c", "cnt": 2}, {"name": "中加国际快递", "num": "scic", "cnt": 2},
                    {"name": "塞尔维亚(PE Post of Serbia)", "num": "serbia", "cnt": 2},
                    {"name": "上大物流", "num": "shangda", "cnt": 2}, {"name": "林道国际快递", "num": "shlindao", "cnt": 2},
                    {"name": "斯洛文尼亚(Slovenia Post)", "num": "slovenia", "cnt": 2},
                    {"name": "南非（South African Post Office）", "num": "southafrican", "cnt": 2},
                    {"name": "深圳DPEX", "num": "szdpex", "cnt": 2}, {"name": "明通国际快递", "num": "tnjex", "cnt": 2},
                    {"name": "德国优拜物流", "num": "ubuy", "cnt": 2}, {"name": "欧洲UEX", "num": "uexiex", "cnt": 2},
                    {"name": "乌克兰邮政包裹", "num": "ukrpostcn", "cnt": 2}, {"name": "西邮寄", "num": "xipost", "cnt": 2},
                    {"name": "银捷速递", "num": "yinjiesudi", "cnt": 2}, {"name": "远盾物流", "num": "yuandun", "cnt": 2},
                    {"name": "德国云快递", "num": "yunexpress", "cnt": 2}, {"name": "转瞬达集运", "num": "zsda56", "cnt": 2},
                    {"name": "明辉物流", "num": "zsmhwl", "cnt": 2}, {"name": "明大快递", "num": "adaexpress", "cnt": 1},
                    {"name": "全程快递", "num": "agopost", "cnt": 1}, {"name": "德方物流", "num": "ahdf", "cnt": 1},
                    {"name": "航空快递", "num": "airgtc", "cnt": 1}, {"name": "AUV国际快递", "num": "auvexpress", "cnt": 1},
                    {"name": "宝通快递", "num": "baotongkd", "cnt": 1},
                    {"name": "喀麦隆(CAMPOST)", "num": "cameroon", "cnt": 1},
                    {"name": "城通物流", "num": "chengtong", "cnt": 1}, {"name": "同舟行物流", "num": "chinatzx", "cnt": 1},
                    {"name": "乌拉圭（Correo Uruguayo）", "num": "correo", "cnt": 1},
                    {"name": "店通快递", "num": "diantongkuaidi", "cnt": 1},
                    {"name": "易达快运", "num": "edaeuexpress", "cnt": 1},
                    {"name": "澳州顺风快递", "num": "emms", "cnt": 1}, {"name": "俄顺达", "num": "eshunda", "cnt": 1},
                    {"name": "EU-EXPRESS", "num": "euexpress", "cnt": 1},
                    {"name": "飞远配送", "num": "feiyuanvipshop", "cnt": 1},
                    {"name": "GHT物流", "num": "ghtexpress", "cnt": 1},
                    {"name": "GSM", "num": "gsm", "cnt": 1}, {"name": "万通快递", "num": "gswtkd", "cnt": 1},
                    {"name": "广通速递", "num": "gtongsudi", "cnt": 1}, {"name": "汉邦国际速递", "num": "handboy", "cnt": 1},
                    {"name": "上海航瑞货运", "num": "hangrui", "cnt": 1}, {"name": "黑狗物流", "num": "higo", "cnt": 1},
                    {"name": "河南全速通", "num": "hnqst", "cnt": 1}, {"name": "红远物流", "num": "hongywl", "cnt": 1},
                    {"name": "环球通达 ", "num": "hqtd", "cnt": 1}, {"name": "华企快运", "num": "huaqikuaiyun", "cnt": 1},
                    {"name": "户通物流", "num": "hutongwuliu", "cnt": 1}, {"name": "大达物流", "num": "idada", "cnt": 1},
                    {"name": "伊朗（Iran Post）", "num": "iran", "cnt": 1},
                    {"name": "以色列(Israel Post)", "num": "israelpost", "cnt": 1},
                    {"name": "jcex", "num": "jcex", "cnt": 1},
                    {"name": "吉祥邮（澳洲）", "num": "jixiangyouau", "cnt": 1},
                    {"name": "骏绅物流", "num": "jsexpress", "cnt": 1},
                    {"name": "KCS", "num": "kcs", "cnt": 1},
                    {"name": "拉脱维亚(Latvijas Pasts)", "num": "latvia", "cnt": 1},
                    {"name": "优能物流", "num": "mantoo", "cnt": 1},
                    {"name": "墨西哥（Correos de Mexico）", "num": "mexico", "cnt": 1},
                    {"name": "银河物流", "num": "milkyway", "cnt": 1}, {"name": "民邦速递", "num": "minbangsudi", "cnt": 1},
                    {"name": "明亮物流", "num": "mingliangwuliu", "cnt": 1}, {"name": "MRW", "num": "mrw", "cnt": 1},
                    {"name": "Newgistics", "num": "newgistics", "cnt": 1}, {"name": "诺尔国际物流", "num": "nuoer", "cnt": 1},
                    {"name": "丹麦(Post Denmark)", "num": "postdanmarken", "cnt": 1},
                    {"name": "土耳其", "num": "ptt", "cnt": 1},
                    {"name": "全晨快递", "num": "quanchenkuaidi", "cnt": 1}, {"name": "全际通", "num": "quanjitong", "cnt": 1},
                    {"name": "全信通快递", "num": "quanxintong", "cnt": 1}, {"name": "瑞丰速递", "num": "rfsd", "cnt": 1},
                    {"name": "日昱物流", "num": "riyuwuliu", "cnt": 1}, {"name": "日日通国际", "num": "rrthk", "cnt": 1},
                    {"name": "Safexpress", "num": "safexpress", "cnt": 1}, {"name": "澳丰速递", "num": "sfau", "cnt": 1},
                    {"name": "顺士达速运", "num": "shunshid", "cnt": 1}, {"name": "信联通", "num": "sinatone", "cnt": 1},
                    {"name": "SkyNet Malaysia", "num": "skynetmalaysia", "cnt": 1},
                    {"name": "TD Cargo", "num": "tdcargo", "cnt": 1},
                    {"name": "Toll Priority(Toll Online)", "num": "tollpriority", "cnt": 1},
                    {"name": "乌克兰小包、大包(UkrPoshta)", "num": "ukraine", "cnt": 1},
                    {"name": "凡客配送（作废）", "num": "vancl", "cnt": 1},
                    {"name": "万庚国际速递", "num": "vangenexpress", "cnt": 1},
                    {"name": "万博快递", "num": "wanboex", "cnt": 1}, {"name": "渥途国际速运", "num": "wotu", "cnt": 1},
                    {"name": "WTD海外通", "num": "wtdex", "cnt": 1}, {"name": "温通物流", "num": "wto56kj", "cnt": 1},
                    {"name": "51跨境通", "num": "wykjt", "cnt": 1}, {"name": "鑫世锐达", "num": "xsrd", "cnt": 1},
                    {"name": "宜送物流", "num": "yiex", "cnt": 1}, {"name": "艺凡快递", "num": "yifankd", "cnt": 1},
                    {"name": "邮来速递", "num": "youlai", "cnt": 1}, {"name": "邮客全球速递", "num": "yyox", "cnt": 1},
                    {"name": "中外运速递", "num": "zhongwaiyun", "cnt": 1}, {"name": "忠信达", "num": "zhongxinda", "cnt": 1},
                    {"name": "泰捷达国际物流", "num": "ztjieda", "cnt": 1}, {"name": "五六快运", "num": "56kuaiyun", "cnt": 0},
                    {"name": "A2U速递", "num": "a2u", "cnt": 0}, {"name": "ABF", "num": "abf", "cnt": 0},
                    {"name": "安达易国际速递", "num": "adiexpress", "cnt": 0}, {"name": "澳多多国际速递", "num": "adodoxm", "cnt": 0},
                    {"name": "ADP国际快递", "num": "adp", "cnt": 0},
                    {"name": "阿富汗(Afghan Post)", "num": "afghan", "cnt": 0},
                    {"name": "AFL", "num": "afl", "cnt": 0}, {"name": "卡邦配送", "num": "ahkbps", "cnt": 0},
                    {"name": "无忧物流", "num": "aliexpress", "cnt": 0},
                    {"name": "亚马逊中国订单", "num": "amazoncnorder", "cnt": 0},
                    {"name": "amazon-国内订单", "num": "amcnorder", "cnt": 0},
                    {"name": "amazon-国际订单", "num": "amusorder", "cnt": 0},
                    {"name": "AOL澳通速递", "num": "aolau", "cnt": 0},
                    {"name": "澳速物流", "num": "aosu", "cnt": 0},
                    {"name": "apgecommerce", "num": "apgecommerce", "cnt": 0},
                    {"name": "美国汉邦快递", "num": "aplus100", "cnt": 0}, {"name": "Aplus物流", "num": "aplusex", "cnt": 0},
                    {"name": "艾瑞斯远", "num": "ariesfar", "cnt": 0},
                    {"name": "阿鲁巴[荷兰]（Post Aruba）", "num": "aruba", "cnt": 0}, {"name": "澳货通", "num": "auex", "cnt": 0},
                    {"name": "澳世速递", "num": "aus", "cnt": 0}, {"name": "澳新物流", "num": "axexpress", "cnt": 0},
                    {"name": "阿塞拜疆EMS(EMS AzerExpressPost)", "num": "azerbaijan", "cnt": 0},
                    {"name": "巴林(Bahrain Post)", "num": "bahrain", "cnt": 0},
                    {"name": "孟加拉国(EMS)", "num": "bangladesh", "cnt": 0},
                    {"name": "报通快递", "num": "baoxianda", "cnt": 0},
                    {"name": "巴巴多斯(Barbados Post)", "num": "barbados", "cnt": 0},
                    {"name": "伯利兹(Belize Postal)", "num": "belize", "cnt": 0},
                    {"name": "白俄罗斯(Belpochta)", "num": "belpost", "cnt": 0},
                    {"name": "笨鸟国际", "num": "benniao", "cnt": 0},
                    {"name": "飛斯特", "num": "bester", "cnt": 0}, {"name": "邦工快运", "num": "bgky100", "cnt": 0},
                    {"name": "BHT", "num": "bht", "cnt": 0}, {"name": "彪记快递", "num": "biaojikuaidi", "cnt": 0},
                    {"name": "青云物流", "num": "bjqywl", "cnt": 0}, {"name": "鑫锐达", "num": "bjxsrd", "cnt": 0},
                    {"name": "标杆物流", "num": "bmlchina", "cnt": 0},
                    {"name": "波黑(JP BH Posta)", "num": "bohei", "cnt": 0},
                    {"name": "玻利维亚", "num": "bolivia", "cnt": 0}, {"name": "BorderGuru", "num": "borderguru", "cnt": 0},
                    {"name": "博茨瓦纳", "num": "botspost", "cnt": 0}, {"name": "速方(Sufast)", "num": "bphchina", "cnt": 0},
                    {"name": "文莱(Brunei Postal)", "num": "brunei", "cnt": 0}, {"name": "展勤快递", "num": "byht", "cnt": 0},
                    {"name": "新喀里多尼亚[法国](New Caledonia)", "num": "caledonia", "cnt": 0},
                    {"name": "柬埔寨(Cambodia Post)", "num": "cambodia", "cnt": 0},
                    {"name": "到了港", "num": "camekong", "cnt": 0}, {"name": "仓鼠快递", "num": "cangspeed", "cnt": 0},
                    {"name": "能装能送", "num": "canhold", "cnt": 0}, {"name": "Canpar", "num": "canpar", "cnt": 0},
                    {"name": "加拿大邮政", "num": "canpostfr", "cnt": 0}, {"name": "卢森堡航空", "num": "cargolux", "cnt": 0},
                    {"name": "钏博物流", "num": "cbo56", "cnt": 0}, {"name": "CCES/国通快递", "num": "cces", "cnt": 0},
                    {"name": "CDEK", "num": "cdek", "cnt": 0}, {"name": "长宇物流", "num": "changyuwuliu", "cnt": 0},
                    {"name": "成达国际速递", "num": "chengda", "cnt": 0}, {"name": "城际快递", "num": "chengji", "cnt": 0},
                    {"name": "城际速递", "num": "chengjisudi", "cnt": 0},
                    {"name": "智利(Correos Chile)", "num": "chile", "cnt": 0},
                    {"name": "SQK国际速递", "num": "chinasqk", "cnt": 0}, {"name": "嘉荣物流", "num": "chllog", "cnt": 0},
                    {"name": "法国大包、EMS-英文(Chronopost France)", "num": "chronopostfren", "cnt": 0},
                    {"name": "Chronopost Portugal", "num": "chronopostport", "cnt": 0},
                    {"name": "城市映急", "num": "city56", "cnt": 0}, {"name": "大韩通运", "num": "cjkoreaexpress", "cnt": 0},
                    {"name": "城晓国际快递", "num": "ckeex", "cnt": 0}, {"name": "澳通华人物流", "num": "cllexpress", "cnt": 0},
                    {"name": "CE易欧通国际速递", "num": "cloudexpress", "cnt": 0}, {"name": "CL日中速运", "num": "clsp", "cnt": 0},
                    {"name": "CNE", "num": "cnexps", "cnt": 0}, {"name": "速舟物流", "num": "cnspeedster", "cnt": 0},
                    {"name": "CNUP 中联邮", "num": "cnup", "cnt": 0}, {"name": "中国翼", "num": "cnws", "cnt": 0},
                    {"name": "哥伦比亚(4-72 La Red Postal de Colombia)", "num": "colombia", "cnt": 0},
                    {"name": "莫桑比克（Correios de Moçambique）", "num": "correios", "cnt": 0},
                    {"name": "阿根廷(Correo Argentina)", "num": "correoargentino", "cnt": 0},
                    {"name": "哥斯达黎加(Correos de Costa Rica)", "num": "correos", "cnt": 0},
                    {"name": "重庆星程快递", "num": "cqxingcheng", "cnt": 0}, {"name": "新时速物流", "num": "csxss", "cnt": 0},
                    {"name": "布谷鸟速递", "num": "cuckooexpess", "cnt": 0},
                    {"name": "塞浦路斯(Cyprus Post)", "num": "cypruspost", "cnt": 0},
                    {"name": "云南诚中物流", "num": "czwlyn", "cnt": 0}, {"name": "达速物流", "num": "dasu", "cnt": 0},
                    {"name": "德中快递", "num": "decnlh", "cnt": 0}, {"name": "德国八易转运", "num": "deguo8elog", "cnt": 0},
                    {"name": "Deltec Courier", "num": "deltec", "cnt": 0},
                    {"name": "澳行快递", "num": "desworks", "cnt": 0},
                    {"name": "东风快递", "num": "dfkuaidi", "cnt": 0}, {"name": "达方物流", "num": "dfpost", "cnt": 0},
                    {"name": "DHL HK", "num": "dhlhk", "cnt": 0},
                    {"name": "DHL-荷兰（DHL Netherlands）", "num": "dhlnetherlands", "cnt": 0},
                    {"name": "DHL-波兰（DHL Poland）", "num": "dhlpoland", "cnt": 0},
                    {"name": "递五方云仓", "num": "di5pll", "cnt": 0}, {"name": "云南滇驿物流", "num": "dianyi", "cnt": 0},
                    {"name": "叮咚快递", "num": "dingdong", "cnt": 0},
                    {"name": "Direct Link", "num": "directlink", "cnt": 0},
                    {"name": "递四方澳洲", "num": "disifangau", "cnt": 0}, {"name": "递四方美国", "num": "disifangus", "cnt": 0},
                    {"name": "天翔东捷运", "num": "djy56", "cnt": 0}, {"name": "东瀚物流", "num": "donghanwl", "cnt": 0},
                    {"name": "东红物流", "num": "donghong", "cnt": 0},
                    {"name": "DPD Germany", "num": "dpdgermany", "cnt": 0},
                    {"name": "DPD Poland", "num": "dpdpoland", "cnt": 0},
                    {"name": "DTDC India", "num": "dtdcindia", "cnt": 0},
                    {"name": "东方航空物流", "num": "ealceair", "cnt": 0},
                    {"name": "E跨通", "num": "ecallturn", "cnt": 0},
                    {"name": "EC-Firstclass", "num": "ecfirstclass", "cnt": 0},
                    {"name": "ECMS Express", "num": "ecmsglobal", "cnt": 0},
                    {"name": "厄瓜多尔(Correos del Ecuador)", "num": "ecuador", "cnt": 0},
                    {"name": "益递物流", "num": "edlogistics", "cnt": 0}, {"name": "龙象国际物流", "num": "edragon", "cnt": 0},
                    {"name": "埃及（Egypt Post）", "num": "egypt", "cnt": 0},
                    {"name": "艾菲尔国际速递", "num": "eiffel", "cnt": 0},
                    {"name": "易联通达", "num": "el56", "cnt": 0},
                    {"name": "希腊包裹（ELTA Hellenic Post）", "num": "elta", "cnt": 0},
                    {"name": "希腊EMS（ELTA Courier）", "num": "eltahell", "cnt": 0},
                    {"name": "阿联酋(Emirates Post)", "num": "emirates", "cnt": 0},
                    {"name": "高考通知书", "num": "emsluqu", "cnt": 0}, {"name": "南非EMS", "num": "emssouthafrica", "cnt": 0},
                    {"name": "乌克兰EMS(EMS Ukraine)", "num": "emsukraine", "cnt": 0},
                    {"name": "乌克兰EMS-中文(EMS Ukraine)", "num": "emsukrainecn", "cnt": 0},
                    {"name": "英国(大包,EMS)", "num": "england", "cnt": 0}, {"name": "联众国际", "num": "epspost", "cnt": 0},
                    {"name": "东方汇", "num": "est365", "cnt": 0}, {"name": "Estafeta", "num": "estafeta", "cnt": 0},
                    {"name": "Estes", "num": "estes", "cnt": 0}, {"name": "易达国际速递", "num": "eta100", "cnt": 0},
                    {"name": "埃塞俄比亚(Ethiopian postal)", "num": "ethiopia", "cnt": 0},
                    {"name": "中欧国际物流", "num": "eucnrail", "cnt": 0},
                    {"name": "德国 EUC POST", "num": "eucpost", "cnt": 0},
                    {"name": "易邮国际", "num": "euguoji", "cnt": 0}, {"name": "败欧洲", "num": "europe8", "cnt": 0},
                    {"name": "优莎速运", "num": "eusacn", "cnt": 0}, {"name": "7号速递", "num": "express7th", "cnt": 0},
                    {"name": "澳洲新干线快递", "num": "expressplus", "cnt": 0}, {"name": "易转运", "num": "ezhuanyuan", "cnt": 0},
                    {"name": "颿达国际快递-英文", "num": "fandaguoji", "cnt": 0},
                    {"name": "颿达国际快递", "num": "fardarww", "cnt": 0},
                    {"name": "丰客物流", "num": "fecobv", "cnt": 0}, {"name": "FedEx-英国件", "num": "fedexukcn", "cnt": 0},
                    {"name": "FedRoad 联邦转运", "num": "fedroad", "cnt": 0},
                    {"name": "飞狐快递", "num": "feihukuaidi", "cnt": 0},
                    {"name": "飞快达", "num": "feikuaida", "cnt": 0}, {"name": "凤凰快递", "num": "fenghuangkuaidi", "cnt": 0},
                    {"name": "斐济(Fiji Post)", "num": "fiji", "cnt": 0}, {"name": "花瓣转运", "num": "flowerkd", "cnt": 0},
                    {"name": "飞力士物流", "num": "flysman", "cnt": 0}, {"name": "全速快递", "num": "fsexp", "cnt": 0},
                    {"name": "法翔速运", "num": "ftlexpress", "cnt": 0}, {"name": "甘肃安的快递", "num": "gansuandi", "cnt": 0},
                    {"name": "高铁快运", "num": "gaotieex", "cnt": 0}, {"name": "Gati-中文", "num": "gaticn", "cnt": 0},
                    {"name": "Gati-英文", "num": "gatien", "cnt": 0}, {"name": "Gati-KWE", "num": "gatikwe", "cnt": 0},
                    {"name": "安的快递", "num": "gda", "cnt": 0}, {"name": "广东诚通物流", "num": "gdct56", "cnt": 0},
                    {"name": "全网物流", "num": "gdqwwl", "cnt": 0}, {"name": "容智快运", "num": "gdrz58", "cnt": 0},
                    {"name": "新鹏快递", "num": "gdxp", "cnt": 0}, {"name": "GE2D跨境物流", "num": "ge2d", "cnt": 0},
                    {"name": "环创物流", "num": "ghl", "cnt": 0},
                    {"name": "直布罗陀[英国]( Royal Gibraltar Post)", "num": "gibraltar", "cnt": 0},
                    {"name": "冠捷物流 ", "num": "gjwl", "cnt": 0},
                    {"name": "globaltracktrace", "num": "globaltracktrace", "cnt": 0},
                    {"name": "英脉物流", "num": "gml", "cnt": 0}, {"name": "UBI Australia", "num": "gotoubi", "cnt": 0},
                    {"name": "格陵兰[丹麦]（TELE Greenland A/S）", "num": "greenland", "cnt": 0},
                    {"name": "潍鸿", "num": "grivertek", "cnt": 0}, {"name": "哥士传奇速递", "num": "gscq365", "cnt": 0},
                    {"name": "GT国际快运", "num": "gtgogo", "cnt": 0}, {"name": "GTS快递", "num": "gts", "cnt": 0},
                    {"name": "广东通路", "num": "guangdongtonglu", "cnt": 0},
                    {"name": "冠庭国际物流", "num": "guanting", "cnt": 0},
                    {"name": "国送快运", "num": "guosong", "cnt": 0}, {"name": "宏观国际快递", "num": "gvpexpress", "cnt": 0},
                    {"name": "贵州星程快递", "num": "gzxingcheng", "cnt": 0},
                    {"name": "海红for买卖宝", "num": "haihongmmb", "cnt": 0},
                    {"name": "海盟速递", "num": "haimengsudi", "cnt": 0}, {"name": "海米派物流", "num": "haimibuy", "cnt": 0},
                    {"name": "海外环球", "num": "haiwaihuanqiu", "cnt": 0},
                    {"name": "海星桥快递", "num": "haixingqiao", "cnt": 0},
                    {"name": "航宇快递", "num": "hangyu", "cnt": 0}, {"name": "昊盛物流", "num": "haoshengwuliu", "cnt": 0},
                    {"name": "好又快物流", "num": "haoyoukuai", "cnt": 0}, {"name": "开心快递", "num": "happylink", "cnt": 0},
                    {"name": "亚美尼亚(Haypost-Armenian Postal)", "num": "haypost", "cnt": 0},
                    {"name": "恒瑞物流", "num": "hengrui56", "cnt": 0}, {"name": "华瀚快递", "num": "hhair56", "cnt": 0},
                    {"name": "Highsince", "num": "highsince", "cnt": 0}, {"name": "海派通", "num": "hipito", "cnt": 0},
                    {"name": "Hi淘易快递", "num": "hitaoe", "cnt": 0}, {"name": "云邮跨境快递", "num": "hkems", "cnt": 0},
                    {"name": "互联快运", "num": "hlkytj", "cnt": 0}, {"name": "共联配", "num": "hlpgyl", "cnt": 0},
                    {"name": "华美快递", "num": "hmus", "cnt": 0}, {"name": "飞鹰物流", "num": "hnfy", "cnt": 0},
                    {"name": "顺时达物流", "num": "hnssd56", "cnt": 0}, {"name": "居家通", "num": "homexpress", "cnt": 0},
                    {"name": "红背心", "num": "hongbeixin", "cnt": 0}, {"name": "宏捷国际物流", "num": "hongjie", "cnt": 0},
                    {"name": "宏品物流", "num": "hongpinwuliu", "cnt": 0}, {"name": "卓烨快递", "num": "hrbzykd", "cnt": 0},
                    {"name": "高铁速递", "num": "hre", "cnt": 0},
                    {"name": "克罗地亚（Hrvatska Posta）", "num": "hrvatska", "cnt": 0},
                    {"name": "海硕高铁速递", "num": "hsgtsd", "cnt": 0}, {"name": "华通快运", "num": "htongexpress", "cnt": 0},
                    {"name": "华通务达物流", "num": "htwd", "cnt": 0}, {"name": "华达快运", "num": "huada", "cnt": 0},
                    {"name": "环东物流", "num": "huandonglg", "cnt": 0}, {"name": "辉联物流", "num": "huilian", "cnt": 0},
                    {"name": "汇强快递", "num": "huiqiangkuaidi", "cnt": 0}, {"name": "驼峰国际", "num": "humpline", "cnt": 0},
                    {"name": "兰州伙伴物流", "num": "huoban", "cnt": 0}, {"name": "鸿远物流", "num": "hyeship", "cnt": 0},
                    {"name": "上海昊宏国际货物", "num": "hyk", "cnt": 0}, {"name": "华航快递", "num": "hzpl", "cnt": 0},
                    {"name": "途鲜物流", "num": "ibenben", "cnt": 0}, {"name": "爱拜物流", "num": "ibuy8", "cnt": 0},
                    {"name": "iExpress", "num": "iexpress", "cnt": 0}, {"name": "logen路坚", "num": "ilogen", "cnt": 0},
                    {"name": "ILYANG", "num": "ilyang", "cnt": 0}, {"name": "艾姆勒", "num": "imlb2c", "cnt": 0},
                    {"name": "印度尼西亚EMS(Pos Indonesia-EMS)", "num": "indonesia", "cnt": 0},
                    {"name": "多米尼加（INPOSDOM – Instituto Postal Dominicano）", "num": "inposdom", "cnt": 0},
                    {"name": "Interlink Express", "num": "interlink", "cnt": 0},
                    {"name": "Italy SDA", "num": "italysad", "cnt": 0},
                    {"name": "牙买加（Jamaica Post）", "num": "jamaicapost", "cnt": 0},
                    {"name": "JDIEX", "num": "jdiex", "cnt": 0}, {"name": "急递", "num": "jdpplus", "cnt": 0},
                    {"name": "佳家通货运", "num": "jiajiatong56", "cnt": 0},
                    {"name": "佳吉快递", "num": "jiajikuaidi", "cnt": 0},
                    {"name": "极光转运", "num": "jiguang", "cnt": 0}, {"name": "锦程物流", "num": "jinchengwuliu", "cnt": 0},
                    {"name": "金大物流", "num": "jindawuliu", "cnt": 0}, {"name": "劲通快递", "num": "jintongkd", "cnt": 0},
                    {"name": "晋越快递", "num": "jinyuekuaidi", "cnt": 0}, {"name": "九宫物流", "num": "jiugong", "cnt": 0},
                    {"name": "久易快递", "num": "jiuyicn", "cnt": 0}, {"name": "佳捷翔物流", "num": "jjx888", "cnt": 0},
                    {"name": "约旦(Jordan Post)", "num": "jordan", "cnt": 0}, {"name": "聚物物流", "num": "juwu", "cnt": 0},
                    {"name": "聚中大", "num": "juzhongda", "cnt": 0}, {"name": "康力物流", "num": "kangliwuliu", "cnt": 0},
                    {"name": "考拉国际速递", "num": "kaolaexpress", "cnt": 0},
                    {"name": "哈萨克斯坦(Kazpost)", "num": "kazpost", "cnt": 0},
                    {"name": "肯尼亚(POSTA KENYA)", "num": "kenya", "cnt": 0}, {"name": "跨境直邮通", "num": "kjde", "cnt": 0},
                    {"name": "番薯国际货运", "num": "koali", "cnt": 0}, {"name": "韩国邮政韩文", "num": "koreapostkr", "cnt": 0},
                    {"name": "淘韩国际快递", "num": "krtao", "cnt": 0}, {"name": "快8速运", "num": "kuai8", "cnt": 0},
                    {"name": "快达物流", "num": "kuaidawuliu", "cnt": 0}, {"name": "快淘快递", "num": "kuaitao", "cnt": 0},
                    {"name": "四川快优达速递", "num": "kuaiyouda", "cnt": 0}, {"name": "凯信达", "num": "kxda", "cnt": 0},
                    {"name": "吉尔吉斯斯坦(Kyrgyz Post)", "num": "kyrgyzpost", "cnt": 0},
                    {"name": "跨跃国际", "num": "kyue", "cnt": 0}, {"name": "蓝镖快递", "num": "lanbiaokuaidi", "cnt": 0},
                    {"name": "蓝弧快递", "num": "lanhukuaidi", "cnt": 0},
                    {"name": "老挝(Lao Express) ", "num": "lao", "cnt": 0},
                    {"name": "塞内加尔", "num": "laposte", "cnt": 0}, {"name": "林安物流", "num": "lasy56", "cnt": 0},
                    {"name": "立白宝凯物流", "num": "lbbk", "cnt": 0}, {"name": "林道国际快递-英文", "num": "ldxpres", "cnt": 0},
                    {"name": "乐递供应链", "num": "ledii", "cnt": 0}, {"name": "乐捷递", "num": "lejiedi", "cnt": 0},
                    {"name": "云豹国际货运", "num": "leopard", "cnt": 0},
                    {"name": "莱索托(Lesotho Post)", "num": "lesotho", "cnt": 0},
                    {"name": "美联快递", "num": "letseml", "cnt": 0},
                    {"name": "龙枫国际快递", "num": "lfexpress", "cnt": 0}, {"name": "lazada", "num": "lgs", "cnt": 0},
                    {"name": "联邦快递-英文", "num": "lianbangkuaidien", "cnt": 0},
                    {"name": "联运快递", "num": "lianyun", "cnt": 0},
                    {"name": "黎巴嫩(Liban Post)", "num": "libanpost", "cnt": 0},
                    {"name": "Linex", "num": "linex", "cnt": 0},
                    {"name": "丽狮物流", "num": "lishi", "cnt": 0},
                    {"name": "立陶宛（Lietuvos pa?tas）", "num": "lithuania", "cnt": 0},
                    {"name": "小熊物流", "num": "littlebearbear", "cnt": 0}, {"name": "良藤国际速递", "num": "lmfex", "cnt": 0},
                    {"name": "龙飞祥快递", "num": "longfx", "cnt": 0}, {"name": "隆浪快递", "num": "longlangkuaidi", "cnt": 0},
                    {"name": "长风物流", "num": "longvast", "cnt": 0}, {"name": "恒通快递", "num": "lqht", "cnt": 0},
                    {"name": "联通快递", "num": "ltparcel", "cnt": 0}, {"name": "论道国际物流", "num": "lundao", "cnt": 0},
                    {"name": "LWE", "num": "lwe", "cnt": 0},
                    {"name": "马其顿(Macedonian Post)", "num": "macedonia", "cnt": 0},
                    {"name": "ME物流", "num": "macroexpressco", "cnt": 0},
                    {"name": "麦力快递", "num": "mailikuaidi", "cnt": 0},
                    {"name": "迈隆递运", "num": "mailongdy", "cnt": 0},
                    {"name": "马尔代夫(Maldives Post)", "num": "maldives", "cnt": 0},
                    {"name": "马耳他（Malta Post）", "num": "malta", "cnt": 0}, {"name": "芒果速递", "num": "mangguo", "cnt": 0},
                    {"name": "今枫国际快运", "num": "mapleexpress", "cnt": 0}, {"name": "美邦国际快递", "num": "meibang", "cnt": 0},
                    {"name": "美达快递", "num": "meidaexpress", "cnt": 0}, {"name": "美泰物流", "num": "meitai", "cnt": 0},
                    {"name": "Mexico Senda Express", "num": "mexicodenda", "cnt": 0},
                    {"name": "闽盛快递", "num": "minshengkuaidi", "cnt": 0}, {"name": "美龙快递", "num": "mjexp", "cnt": 0},
                    {"name": "摩尔多瓦(Posta Moldovei)", "num": "moldova", "cnt": 0},
                    {"name": "蒙古国(Mongol Post) ", "num": "mongolpost", "cnt": 0},
                    {"name": "黑山(Posta Crne Gore)", "num": "montenegro", "cnt": 0},
                    {"name": "摩洛哥 ( Morocco Post )", "num": "morocco", "cnt": 0},
                    {"name": "魔速达", "num": "mosuda", "cnt": 0},
                    {"name": "Mexico Multipack", "num": "multipack", "cnt": 0},
                    {"name": "中俄速通（淼信）", "num": "mxe56", "cnt": 0},
                    {"name": "纳米比亚(NamPost)", "num": "namibia", "cnt": 0},
                    {"name": "红马速递", "num": "nedahm", "cnt": 0},
                    {"name": "尼泊尔（Nepal Postal Services）", "num": "nepalpost", "cnt": 0},
                    {"name": "尼日利亚(Nigerian Postal)", "num": "nigerianpost", "cnt": 0},
                    {"name": "浩博物流", "num": "njhaobo", "cnt": 0}, {"name": "NLE", "num": "nle", "cnt": 0},
                    {"name": "亚欧专线", "num": "nlebv", "cnt": 0}, {"name": "华赫物流", "num": "nmhuahe", "cnt": 0},
                    {"name": "Nova Poshta", "num": "novaposhta", "cnt": 0},
                    {"name": "偌亚奥国际快递", "num": "nuoyaao", "cnt": 0},
                    {"name": "新西兰中通", "num": "nzzto", "cnt": 0},
                    {"name": "OBOR Express", "num": "oborexpress", "cnt": 0},
                    {"name": "OCA Argentina", "num": "ocaargen", "cnt": 0},
                    {"name": "阿曼(Oman Post)", "num": "oman", "cnt": 0},
                    {"name": "爱沙尼亚(Eesti Post)", "num": "omniva", "cnt": 0}, {"name": "OPEK", "num": "opek", "cnt": 0},
                    {"name": "中欧快运", "num": "otobv", "cnt": 0}, {"name": "波音速递", "num": "overseaex", "cnt": 0},
                    {"name": "巴基斯坦(Pakistan Post)", "num": "pakistan", "cnt": 0},
                    {"name": "巴拉圭(Correo Paraguayo)", "num": "paraguay", "cnt": 0},
                    {"name": "英国邮政大包EMS", "num": "parcelforcecn", "cnt": 0},
                    {"name": "全球速递", "num": "pdstow", "cnt": 0},
                    {"name": "派尔快递", "num": "peex", "cnt": 0}, {"name": "配思货运", "num": "peisihuoyunkuaidi", "cnt": 0},
                    {"name": "陪行物流", "num": "peixingwuliu", "cnt": 0}, {"name": "鹏程快递", "num": "pengcheng", "cnt": 0},
                    {"name": "秘鲁(SERPOST)", "num": "peru", "cnt": 0},
                    {"name": "菲律宾（Philippine Postal）", "num": "phlpost", "cnt": 0},
                    {"name": "品速心达快递", "num": "pinsuxinda", "cnt": 0},
                    {"name": "品信快递", "num": "pinxinkuaidi", "cnt": 0},
                    {"name": "先锋国际快递", "num": "pioneer", "cnt": 0},
                    {"name": "北极星快运", "num": "polarisexpress", "cnt": 0},
                    {"name": "葡萄牙（Portugal CTT）", "num": "portugalctt", "cnt": 0},
                    {"name": "Portugal Seur", "num": "portugalseur", "cnt": 0},
                    {"name": "PostNord(Posten AB)", "num": "postenab", "cnt": 0},
                    {"name": "挪威（Posten Norge）", "num": "postennorge", "cnt": 0},
                    {"name": "巴布亚新几内亚(PNG Post)", "num": "postpng", "cnt": 0},
                    {"name": "Purolator", "num": "purolator", "cnt": 0}, {"name": "急顺通", "num": "pzhjst", "cnt": 0},
                    {"name": "ANTS EXPRESS", "num": "qdants", "cnt": 0}, {"name": "雪域快递", "num": "qhxykd", "cnt": 0},
                    {"name": "千里速递", "num": "qianli", "cnt": 0}, {"name": "卡塔尔（Qatar Post）", "num": "qpost", "cnt": 0},
                    {"name": "千顺快递", "num": "qskdyxgs", "cnt": 0}, {"name": "全川物流", "num": "quanchuan56", "cnt": 0},
                    {"name": "全速通", "num": "quansutong", "cnt": 0}, {"name": "Quantium", "num": "quantium", "cnt": 0},
                    {"name": "全之鑫物流", "num": "qzx56", "cnt": 0}, {"name": "Red Express", "num": "redexpress", "cnt": 0},
                    {"name": "睿和泰速运", "num": "rhtexpress", "cnt": 0}, {"name": "荣庆物流", "num": "rokin", "cnt": 0},
                    {"name": "罗马尼亚（Posta Romanian）", "num": "romanian", "cnt": 0},
                    {"name": "rpx", "num": "rpx", "cnt": 0},
                    {"name": "捷网俄全通", "num": "ruexp", "cnt": 0}, {"name": "瑞达国际速递", "num": "ruidaex", "cnt": 0},
                    {"name": "全时速运", "num": "runhengfeng", "cnt": 0},
                    {"name": "卢旺达(Rwanda i-posita)", "num": "rwanda", "cnt": 0},
                    {"name": "日益通速递", "num": "rytsd", "cnt": 0}, {"name": "赛澳递", "num": "saiaodi", "cnt": 0},
                    {"name": "萨摩亚(Samoa Post)", "num": "samoa", "cnt": 0},
                    {"name": "三盛快递", "num": "sanshengco", "cnt": 0},
                    {"name": "海信物流", "num": "savor", "cnt": 0}, {"name": "四川星程快递", "num": "scxingcheng", "cnt": 0},
                    {"name": "速呈", "num": "sczpds", "cnt": 0}, {"name": "首达速运", "num": "sdsy888", "cnt": 0},
                    {"name": "优配速运", "num": "sdyoupei", "cnt": 0},
                    {"name": "Selektvracht", "num": "selektvracht", "cnt": 0},
                    {"name": "International Seur", "num": "seur", "cnt": 0},
                    {"name": "十方通物流", "num": "sfift", "cnt": 0},
                    {"name": "圣飞捷快递", "num": "sfjhd", "cnt": 0}, {"name": "曹操到", "num": "sfpost", "cnt": 0},
                    {"name": "衫达快运", "num": "shanda56", "cnt": 0},
                    {"name": "上海快通", "num": "shanghaikuaitong", "cnt": 0},
                    {"name": "上海无疆for买卖宝", "num": "shanghaiwujiangmmb", "cnt": 0},
                    {"name": "捎客物流", "num": "shaoke", "cnt": 0}, {"name": "盛通快递", "num": "shengtongscm", "cnt": 0},
                    {"name": "神马快递", "num": "shenma", "cnt": 0}, {"name": "阳光快递", "num": "shiningexpress", "cnt": 0},
                    {"name": "王牌快递", "num": "shipbyace", "cnt": 0}, {"name": "苏豪快递", "num": "shipsoho", "cnt": 0},
                    {"name": "世运快递", "num": "shiyunkuaidi", "cnt": 0}, {"name": "wish邮", "num": "shpostwish", "cnt": 0},
                    {"name": "顺邦国际物流", "num": "shunbang", "cnt": 0}, {"name": "四海快递", "num": "sihaiet", "cnt": 0},
                    {"name": "四海捷运", "num": "sihiexpress", "cnt": 0}, {"name": "中外运空运", "num": "sinoairinex", "cnt": 0},
                    {"name": "中外运速递-中文", "num": "sinoex", "cnt": 0}, {"name": "Siodemka", "num": "siodemka", "cnt": 0},
                    {"name": "易普递", "num": "sixroad", "cnt": 0}, {"name": "skynet", "num": "skynet", "cnt": 0},
                    {"name": "skynetworldwide", "num": "skynetworldwide", "cnt": 0},
                    {"name": "荷兰Sky Post", "num": "skypost", "cnt": 0},
                    {"name": "斯洛伐克(Slovenská Posta)", "num": "slovak", "cnt": 0},
                    {"name": "斯里兰卡(Sri Lanka Post)", "num": "slpost", "cnt": 0},
                    {"name": "嗖一下同城快递", "num": "sofast56", "cnt": 0}, {"name": "行必达", "num": "speeda", "cnt": 0},
                    {"name": "申必达", "num": "speedoex", "cnt": 0}, {"name": "首通快运", "num": "staky", "cnt": 0},
                    {"name": "星速递", "num": "starex", "cnt": 0}, {"name": "星运快递", "num": "staryvr", "cnt": 0},
                    {"name": "速豹", "num": "subaoex", "cnt": 0}, {"name": "速呈宅配", "num": "sucheng", "cnt": 0},
                    {"name": "特急便物流", "num": "sucmj", "cnt": 0}, {"name": "苏丹（Sudapost）", "num": "sudapost", "cnt": 0},
                    {"name": "穗佳物流", "num": "suijiawuliu", "cnt": 0}, {"name": "郑州速捷", "num": "sujievip", "cnt": 0},
                    {"name": "速配欧翼", "num": "superoz", "cnt": 0}, {"name": "速品快递", "num": "supinexpress", "cnt": 0},
                    {"name": "Sureline冠泰", "num": "sureline", "cnt": 0}, {"name": "天越物流", "num": "surpassgo", "cnt": 0},
                    {"name": "狮爱高铁物流", "num": "sycawl", "cnt": 0}, {"name": "天美快递", "num": "taimek", "cnt": 0},
                    {"name": "坦桑尼亚(Tanzania Posts)", "num": "tanzania", "cnt": 0},
                    {"name": "TCI XPS", "num": "tcixps", "cnt": 0}, {"name": "天翔快递", "num": "tianxiang", "cnt": 0},
                    {"name": "天纵物流", "num": "tianzong", "cnt": 0}, {"name": "万家通快递", "num": "timedg", "cnt": 0},
                    {"name": "天联快运", "num": "tlky", "cnt": 0}, {"name": "TNT Italy", "num": "tntitaly", "cnt": 0},
                    {"name": "TNT Post", "num": "tntpostcn", "cnt": 0}, {"name": "TNY物流", "num": "tny", "cnt": 0},
                    {"name": "Toll", "num": "toll", "cnt": 0}, {"name": "通达兴物流", "num": "tongdaxing", "cnt": 0},
                    {"name": "通和天下", "num": "tonghetianxia", "cnt": 0}, {"name": "顶世国际物流", "num": "topshey", "cnt": 0},
                    {"name": "天使物流云", "num": "tswlcloud", "cnt": 0},
                    {"name": "突尼斯EMS(Rapid-Poste)", "num": "tunisia", "cnt": 0},
                    {"name": "海龟国际快递", "num": "turtle", "cnt": 0}, {"name": "UEX国际物流", "num": "uex", "cnt": 0},
                    {"name": "乌干达(Posta Uganda)", "num": "uganda", "cnt": 0},
                    {"name": "邮鸽速运", "num": "ugoexpress", "cnt": 0},
                    {"name": "乌克兰小包、大包(UkrPost)", "num": "ukrpost", "cnt": 0},
                    {"name": "UPS Freight", "num": "upsfreight", "cnt": 0},
                    {"name": "UPS Mail Innovations", "num": "upsmailinno", "cnt": 0},
                    {"name": "UTAO优到", "num": "utaoscm", "cnt": 0},
                    {"name": "瓦努阿图(Vanuatu Post)", "num": "vanuatu", "cnt": 0},
                    {"name": "越中国际物流", "num": "vctrans", "cnt": 0},
                    {"name": "越南小包(Vietnam Posts)", "num": "vietnam", "cnt": 0},
                    {"name": "鹰运国际速递", "num": "vipexpress", "cnt": 0}, {"name": "维普恩物流", "num": "vps", "cnt": 0},
                    {"name": "宁夏万家通", "num": "wanjiatong", "cnt": 0}, {"name": "万达美", "num": "wdm", "cnt": 0},
                    {"name": "未来明天快递", "num": "weilaimingtian", "cnt": 0},
                    {"name": "文捷航空", "num": "wenjiesudi", "cnt": 0},
                    {"name": "万邑通", "num": "winit", "cnt": 0}, {"name": "凡仕特物流", "num": "wlfast", "cnt": 0},
                    {"name": "伍圆速递", "num": "wuyuansudi", "cnt": 0}, {"name": "万运国际快递", "num": "wygj168", "cnt": 0},
                    {"name": "微转运", "num": "wzhaunyun", "cnt": 0}, {"name": "西安胜峰", "num": "xaetc", "cnt": 0},
                    {"name": "国晶物流", "num": "xdshipping", "cnt": 0},
                    {"name": "西安城联速递", "num": "xianchengliansudi", "cnt": 0},
                    {"name": "先锋快递", "num": "xianfeng", "cnt": 0},
                    {"name": "湘达物流", "num": "xiangdawuliu", "cnt": 0}, {"name": "翔腾物流", "num": "xiangteng", "cnt": 0},
                    {"name": "小C海淘", "num": "xiaocex", "cnt": 0}, {"name": "西安喜来快递", "num": "xilaikd", "cnt": 0},
                    {"name": "新元快递", "num": "xingyuankuaidi", "cnt": 0},
                    {"name": "信天捷快递", "num": "xintianjie", "cnt": 0},
                    {"name": "鑫通宝物流", "num": "xtb", "cnt": 0}, {"name": "鑫远东速运", "num": "xyd666", "cnt": 0},
                    {"name": "亚马逊中国", "num": "yamaxunwuliu", "cnt": 0}, {"name": "乾坤物流", "num": "yatexpress", "cnt": 0},
                    {"name": "一辉物流", "num": "yatfai", "cnt": 0}, {"name": "YCG物流", "num": "ycgglobal", "cnt": 0},
                    {"name": "也门(Yemen Post)", "num": "yemen", "cnt": 0}, {"name": "驭丰速运", "num": "yfsuyun", "cnt": 0},
                    {"name": "宇航通物流", "num": "yhtlogistics", "cnt": 0},
                    {"name": "一邦速递", "num": "yibangwuliu", "cnt": 0},
                    {"name": "驿递汇速递", "num": "yidihui", "cnt": 0}, {"name": "易航物流", "num": "yihangmall", "cnt": 0},
                    {"name": "亿领速运", "num": "yilingsuyun", "cnt": 0}, {"name": "音素快运", "num": "yinsu", "cnt": 0},
                    {"name": "一柒国际物流", "num": "yiqiguojiwuliu", "cnt": 0},
                    {"name": "亿顺航", "num": "yishunhang", "cnt": 0},
                    {"name": "宜送", "num": "yisong", "cnt": 0}, {"name": "易邮速运", "num": "yiyou", "cnt": 0},
                    {"name": "一正达速运", "num": "yizhengdasuyun", "cnt": 0}, {"name": "洋口岸", "num": "ykouan", "cnt": 0},
                    {"name": "永邦国际物流", "num": "yongbangwuliu", "cnt": 0},
                    {"name": "永旺达快递", "num": "yongwangda", "cnt": 0},
                    {"name": "友家速递", "num": "youjia", "cnt": 0}, {"name": "雅澳物流", "num": "yourscm", "cnt": 0},
                    {"name": "壹品速递", "num": "ypsd", "cnt": 0}, {"name": "运通快运", "num": "ytky168", "cnt": 0},
                    {"name": "远航国际快运", "num": "yuanhhk", "cnt": 0},
                    {"name": "元智捷诚", "num": "yuanzhijiecheng", "cnt": 0},
                    {"name": "越丰物流", "num": "yuefengwuliu", "cnt": 0},
                    {"name": "粤中国际货运代理（上海）有限公司", "num": "yuezhongsh", "cnt": 0},
                    {"name": "御风速运", "num": "yufeng", "cnt": 0}, {"name": "煜嘉物流", "num": "yujiawuliu", "cnt": 0},
                    {"name": "运东西", "num": "yundx", "cnt": 0}, {"name": "韵丰物流", "num": "yunfeng56", "cnt": 0},
                    {"name": "燕文物流", "num": "yw56", "cnt": 0}, {"name": "远为快递", "num": "ywexpress", "cnt": 0},
                    {"name": "西安运逸快递", "num": "yyexp", "cnt": 0}, {"name": "直德邮", "num": "zdepost", "cnt": 0},
                    {"name": "三三国际物流", "num": "zenzen", "cnt": 0}, {"name": "珠峰速运", "num": "zf365", "cnt": 0},
                    {"name": "一站通快递", "num": "zgyzt", "cnt": 0}, {"name": "众辉达物流", "num": "zhdwl", "cnt": 0},
                    {"name": "至诚通达快递", "num": "zhichengtongda", "cnt": 0},
                    {"name": "志腾物流", "num": "zhitengwuliu", "cnt": 0},
                    {"name": "中技物流", "num": "zhongjiwuliu", "cnt": 0},
                    {"name": "中睿速递", "num": "zhongruisudi", "cnt": 0}, {"name": "众派速递", "num": "zhpex", "cnt": 0},
                    {"name": "卓实快运", "num": "zhuoshikuaiyun", "cnt": 0}, {"name": "青旅物流", "num": "zqlwl", "cnt": 0},
                    {"name": "准实快运", "num": "zsky123", "cnt": 0}, {"name": "智通物流", "num": "ztong", "cnt": 0}]

    def __init__(self, path, activeCall):
        self.ActiveCall = activeCall
        self.excelObject = Workbook()  # 打开一个将写的文件
        self.path = path

        self.onePrice = 0.08

        self.logsSheetTitle = ['订阅日期', '编码', '快递公司编码', '运单号', '金额', '快递公司名称']
        self.logsSheetRow = 1

        self.sumSheetTitle = ['行标签', '运单号', '金额']
        self.sumSheetRow = 1

        self.expressSheetTitle = ['公司编码', '公司名称']
        self.expressSheetRow = 1

        self.createFlag = False

        self.process(5, '构造表单标题成功')

        self.main()

    def main(self):
        with open(self.path, 'r', encoding='utf-8') as file:
            # 初始化算集合的值{公司编号:[公司名称,累计数量,后边加上累计费用]}
            arr = {}
            file = file.readlines()
            # row = 1
            self.process(7, '打开文件')
            for key in file:
                (time, KDnumber, epxressCompany, expressNumber, status) = key.split("|")
                epxressCompanyName = self.getExpressNameByKey(epxressCompany)
                # 第一次创建表
                if not self.createFlag:
                    self.createSheet(KDnumber)
                if epxressCompany in arr.keys():
                    # 存在的时候 先读取到他的key给加上1
                    arr.update({epxressCompany: [epxressCompanyName, arr[epxressCompany][1] + 1]})
                else:
                    # 不存在写入一个
                    arr.update({epxressCompany: [epxressCompanyName, 1]})
                # 写入logo表
                data = [time, KDnumber, epxressCompany, expressNumber, self.onePrice, epxressCompanyName]
                self.write(self.logsSheet, 'logsSheetRow', data)
                # break
            self.process(65, 'log表写入成功')
            # 计算价格
            sumPrice = 0
            sumQuantity = 0
            for key in arr.keys():
                price = self.getPrice(arr[key][1])
                sumPrice += price
                sumQuantity += arr[key][1]
                arr.update({key: [arr[key][0], arr[key][1], price]})
                # 计算价格
            arr = sorted(arr.values(), key=lambda x: x[1], reverse=True)
            sumPrice = round(sumPrice, 2)

            # 每个快递累计
            for key in arr:
                self.write(self.SummarySheet, 'sumSheetRow', [key[0], key[1], key[2]])
            # 总计
            self.write(self.SummarySheet, 'sumSheetRow', [str('总计'), str(sumQuantity), str(sumPrice)])
            self.process(95, '汇总表写入成功')

            saveExcel = os.path.split(self.path)[0] + "/" + self.fileName + ".xlsx"
            self.excelObject.save(saveExcel)
            self.process(100, '保存成功')
            self.ActiveCall.finish.emit(str(saveExcel))

    def write(self, object, row, data, isHeader=False):
        """
        写入某行
        :param object:需要写的表的对象
        :param row: 行数
        :param data: 写入的list数据
        :return:
        """
        if row == 'logsSheetRow':
            self.writeToCol(object, self.logsSheetRow, data, isHeader)
            self.logsSheetRow += 1
        elif row == 'expressSheetRow':
            # row = self.expressSheetRow
            self.writeToCol(object, self.expressSheetRow, data, isHeader)
            self.expressSheetRow += 1
        else:
            self.writeToCol(object, self.sumSheetRow, data, isHeader)
            self.sumSheetRow += 1

    def writeToCol(self, object, row, data, isHeader):
        col = 1
        for s in data:  # 再循环里面list的值，每一列
            object.cell(row, col).value = str(s)  # 写文件
            if isHeader == True:
                object.cell(row, col).font = Font(bold=True)
            col += 1

    def getPrice(self, number):
        """
        获取价格
        :param number:
        :return:
        """
        return round(number * self.onePrice, 2)

    def createSheet(self, KDnumber):
        """
        创建表并且写入标题
        :param KDnumber:
        :return:
        """
        now_time = datetime.now()
        self.fileName = now_time.strftime(KDnumber + '.%Y-%m')
        self.logsSheet = self.excelObject.create_sheet(self.fileName, index=0)
        self.write(self.logsSheet, 'logsSheetRow', self.logsSheetTitle, True)

        self.ExpressSheet = self.excelObject.create_sheet('快递编码', index=0)
        self.write(self.ExpressSheet, 'expressSheetRow', self.expressSheetTitle, True)

        self.SummarySheet = self.excelObject.create_sheet('汇总', index=0)
        self.write(self.SummarySheet, 'sumSheetRow', self.sumSheetTitle, True)

        self.process(15, 'sheet构造成功')
        self.createFlag = True

        self.insertDataForExpress()
        self.process(35, '快递公司对照表写入成功')

        pass

    def insertDataForExpress(self):
        """
        将快递表写入
        :param obj:
        :return:
        """
        for a in miaKd100Pricce.EXPRESS_INFO:
            self.write(self.ExpressSheet, 'expressSheetRow', [a['num'], a['name']])

    def getExpressNameByKey(self, key):
        """
        获取名称
        :param key:
        :return:
        """
        for a in miaKd100Pricce.EXPRESS_INFO:
            if a['num'] == key:
                return a['name']
        return False

    def process(self, value, text):
        """
        写入日志
        :param value:
        :param text:
        :return:
        """
        self.ActiveCall.log.emit(str(text))
        self.ActiveCall.dealEventOfGUI.emit(int(value))


class UI_FROM(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(632, 476)
        self.centralwidget = QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.verticalLayout_2 = QVBoxLayout(self.centralwidget)
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.verticalLayout = QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")
        self.textBrowser = QTextBrowser(self.centralwidget)
        self.textBrowser.setObjectName("textBrowser")
        self.verticalLayout.addWidget(self.textBrowser)
        self.progressBar = QProgressBar(self.centralwidget)
        self.progressBar.setProperty("value", 0)
        self.progressBar.setObjectName("progressBar")
        self.verticalLayout.addWidget(self.progressBar)
        self.horizontalLayout = QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.pushButton_2 = QPushButton(self.centralwidget)
        self.pushButton_2.setObjectName("pushButton_2")
        self.horizontalLayout.addWidget(self.pushButton_2)
        self.pushButton = QPushButton(self.centralwidget)
        self.pushButton.setObjectName("pushButton")
        self.horizontalLayout.addWidget(self.pushButton)
        self.verticalLayout.addLayout(self.horizontalLayout)
        self.verticalLayout_2.addLayout(self.verticalLayout)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QMenuBar(MainWindow)
        self.menubar.setGeometry(QRect(0, 0, 632, 23))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "快递100日志生成EXCEL报表"))
        import ctypes
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("myappid")
        MainWindow.setWindowIcon(QIcon(':/source/image/logo.ico'))
        self.pushButton_2.setText(_translate("MainWindow", "选择文件"))
        self.pushButton.setText(_translate("MainWindow", "开始"))


class dealEvent(QMainWindow, UI_FROM):
    def __init__(self, parent=None):
        QMainWindow.__init__(self, parent)
        super(UI_FROM, self).__init__()
        self.setupUi(self)
        self.retranslateUi(self)

        self.pushButton.setEnabled(False)
        self.pushButton_2.clicked.connect(self.selectFile)
        self.pushButton.clicked.connect(self.process)

    def selectFile(self):
        file_name = QFileDialog.getOpenFileName(self, '选择文件', '', 'log文件(*.log)')
        if (file_name[0] == ""):
            QMessageBox.information(self, "提示", self.tr("没有选择任何文件!"))
            self.log('取消了选择文件')
            # self.pushButton_2.
            if self.pushButton_2.text() == "":
                self.pushButton.setEnabled(False)
        else:
            self.log('选择了文件' + file_name[0])
            self.pushButton_2.setText(file_name[0])
            self.pushButton.setEnabled(True)

    def process(self):
        if self.pushButton_2.text() == "":
            self.pushButton.setEnabled(False)
            QMessageBox.information(self, "提示", self.tr("没有选择任何文件!"))
        else:
            self.pushButton_2.setEnabled(False)
            self.pushButton.setEnabled(False)
            self.th = WorkThread(filePath=self.pushButton_2.text())
            self.th.log.connect(self.log)
            self.th.dealEventOfGUI.connect(self.dealEventOfGUI)
            self.th.finish.connect(self.finish)
            # 启动线程
            self.th.start()

    def finish(self, path):
        self.pushButton_2.setEnabled(True)
        self.pushButton.setEnabled(True)
        res = QMessageBox.information(self,
                                      "信息",
                                      "任务执行完成，是否打开该文件",
                                      QMessageBox.Yes | QMessageBox.No)
        if 16384 == res:
            os.system(path)

    def log(self, msg):
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        event = "时间：{}，操作：{}。".format(now, msg)
        self.textBrowser.append(event)

    def dealEventOfGUI(self, process):
        self.progressBar.setValue(process)
        if process == 0:
            QMessageBox.information(self,
                                    "异常",
                                    "发生了意外,请检查",
                                    QMessageBox.Yes)

    def closeEvent(self, event):
        reply = QMessageBox.question(self,
                                     '确认退出',
                                     "是否要退出程序？",
                                     QMessageBox.Yes | QMessageBox.No,
                                     QMessageBox.No)
        if reply == QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()


class WorkThread(QThread):
    log = pyqtSignal(str)
    dealEventOfGUI = pyqtSignal(int)
    finish = pyqtSignal(str)
    def __init__(self, filePath, parent=None):
        super(WorkThread, self).__init__(parent)
        self.filePath = filePath

    def run(self):
        try:
            self.log.emit('开始处理文件，选择文件为' + self.filePath)
            self.dealEventOfGUI.emit(1)
            a = miaKd100Pricce(self.filePath, self)
            self.log.emit('文件处理成功,请检查输入目录')
        except Exception as e:
            self.dealEventOfGUI.emit(0)
            self.log.emit('发送意外：' + str(e))


if __name__ == "__main__":
    app = QApplication(sys.argv)  # 初始化app
    splash = QSplashScreen(QPixmap(":/source/image/loading.png"))
    splash.showMessage("loading, please wait!")
    splash.show()
    MainWindow = QMainWindow()  # 创建主窗口
    ui = dealEvent()
    sleep(2)
    ui.show()  # 显示窗口
    splash.finish(ui)
    sys.exit(app.exec_())  # 消息循环结束之后返回0，接着调用sys.exit(0)退出程序
